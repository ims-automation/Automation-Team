import openpyxl
import re
import pandas as ra
import ipaddress
import pdb
import socket
import struct
import os

excel_path = input("Please Enter the Excel file path, full path with itz name and extension also \n ")
def read_write_excel():
	
    w = openpyxl.load_workbook(r"{}".format(excel_path))
    sheet = w['IP Addressing']
    d = {}
    a = []  # List of all Component column items,with their position of occurrence
    b = []  # List of  the position of all the occurrence of the Component column, will take from list 'a'
    for i in range(2,sheet.max_row):
        if type(sheet['B{}'.format(i)].value) == str:
            a.append((sheet['B{}'.format(i)].value,i))

    for al in a:        # Adding the position of all the occurrence of the Component column to list 'b'
        b.append(al[1])

    x = []  # To get the last cell of  column 'Network'
    for i in range(2,sheet.max_row):
        if sheet['D{}'.format(i)].value == None:
            x.append((sheet['D{}'.format(i)].value,i))
            break

    b.append(x[0][1])  # adding the maximum length of sheet to the list 'b'

    for k in range(len(b)-1):   # Creating the the Dict of the Component with their Network value
        c = []
        for j in range(b[k],b[k+1]):
            c.append(sheet['D{}'.format(j)].value)
        d[sheet['B{}'.format(b[k])].value] = c

    f = open("data.txt",'w')
    for e,r in d.items():       # Writing the data to the file
        dig = re.search(r'\d',e).group()
        #pdb.set_trace()
        if int(dig)==0:
            for g in r:
                for m in range(2,x[0][1]):
                    try:
                        if g == 'Internal':
                            for y in a:
                                if e in y:
                                    pt = y[1]
                                    name_ = sheet['A{}'.format(pt)].value
                                    if type(name_) is not str:
                                        for rg in range(pt, 0, -1):
                                            if type(sheet['A{}'.format(rg)].value) == str:
                                                name_ = sheet['A{}'.format(rg)].value
                                                break

                                    if name_.split()[0] in sheet['I{}'.format(m)].value:
                                        try:
                                            f.write('{},{},{},{}\n'.format(e, sheet['I{}'.format(m)].value,
                                                                           (sheet['K{}'.format(m)].value).split('/')[0],
                                                                           (sheet['K{}'.format(m)].value).split('/')[
                                                                               1]))
                                        except AttributeError:
                                            pass
                        else:
                            if g == sheet['I{}'.format(m)].value:
                                try:
                                    f.write('{},{},{},{}\n'.format(e,sheet['I{}'.format(m)].value,(sheet['K{}'.format(m)].value).split('/')[0],(sheet['K{}'.format(m)].value).split('/')[1]))
                                except AttributeError:
                                    pass
                    except TypeError:
                        pass
        else:

            for dg in range(1,int(dig)+1):
                for g in r:
                    for m in range(2,x[0][1]):
                        try:
                            if g == 'Internal':
                                for y in a:
                                    if e in y:
                                        pt = y[1]
                                        name_ = sheet['A{}'.format(pt)].value
                                        if type(name_) is not str:
                                            for rg in range(pt,0,-1):
                                                if type(sheet['A{}'.format(rg)].value) == str:
                                                    name_ = sheet['A{}'.format(rg)].value
                                                    break

                                        if name_.split()[0] in sheet['I{}'.format(m)].value and name_.split()[0] != 'CGF':
                                                try:
                                                    f.write(
                                                        '{}_{},{},{},{}\n'.format(e, dg, sheet['I{}'.format(m)].value,
                                                                                  (sheet['K{}'.format(m)].value).split(
                                                                                      '/')[0],
                                                                                  (sheet['K{}'.format(m)].value).split(
                                                                                      '/')[1]))
                                                except AttributeError:
                                                    pass
                            else:
                                if g == sheet['I{}'.format(m)].value:
                                    try:
                                        f.write('{}_{},{},{},{}\n'.format(e,dg,sheet['I{}'.format(m)].value,(sheet['K{}'.format(m)].value).split('/')[0],(sheet['K{}'.format(m)].value).split('/')[1]))
                                    except AttributeError:
                                        pass

                        except TypeError:
                            pass


    f.close()

    df = ra.read_csv("data.txt",names=('Node Descripton','Network Purpose','Network Address','Sub Net','IP Address','Next Hop'))
    writer = ra.ExcelWriter(r"{}".format(excel_path),engine='openpyxl')
    writer.book = w
    df.to_excel(excel_writer=writer,sheet_name='Master_Mavenir',index=False)
    w.save(r"{}".format(excel_path))

    writer.save()
    #print(d)


def ip_address_modification():  # For IP Address and Next Hop
    index =0
    w = openpyxl.load_workbook(r"{}".format(excel_path))
    sheet = w['Master_Mavenir']
    list_elem = set()

    for i in range(2,sheet.max_row):
        list_elem.add(sheet['B{}'.format(i)].value)


    for j in list_elem:

        for k in range(2,sheet.max_row+1):
            if j == sheet['B{}'.format(k)].value:
                index += 1
                ip = sheet['C{}'.format(k)].value
                m = struct.unpack('!L', socket.inet_aton(ip))[0]    # Converting the Ip from str to int and in the next line adding one digit and assining
                sheet['E{}'.format(k)] = (socket.inet_ntoa(struct.pack('!L', m + index+1)))
                sheet['F{}'.format(k)] = (socket.inet_ntoa(struct.pack('!L', m + 1)))

        index = 0


    w.save(r"{}".format(excel_path))


if __name__ == '__main__':
    read_write_excel()
    ip_address_modification()
    #for cdr in os.listdir(os.getcwd()):
    	#if cdr.endswith('.txt'):
    		#os.unlink(os.path.join(os.getcwd(),cdr))

    print('Successfull Completed')
    input('Enter to exit')
